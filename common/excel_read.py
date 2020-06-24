# -*- coding: utf-8 -*-
# @Time : 2020/5/30 16:28
# @Author : 深圳-烧烤-29期
# @File : excel_read.py

"""读取Excel测试用例数据"""
import openpyxl


class ExcelHandler:
    def __init__(self,path):
        self.path = path
        self.workbook = None

    def open_excel(self):
        workbook = openpyxl.load_workbook(self.path)
        self.workbook = workbook
        return workbook

    def get_sheet(self,sheet_name):
        sheet = self.open_excel()[sheet_name]
        return sheet

    def read_data(self,sheet_name):
        sheet = self.get_sheet(sheet_name)
        data = []
        for row in range(2, sheet.max_row + 1):
            dic = {}
            sub_dic = {}
            for col in range(1, sheet.max_column + 1):
                sub_dic[sheet.cell(1, col).value] = sheet.cell(row, col).value
            dic.update(sub_dic)
            data.append(dic)
        return data

    def write_excel(self,sheet_name, row, col, value):
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row, col).value = value
        self.save_excel()
        self.close_excel()

    def save_excel(self):
        self.workbook.save(self.path)

    def close_excel(self):
        self.workbook.close()


